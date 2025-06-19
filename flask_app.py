import os
import uuid
import zipfile
import shutil
import traceback
from flask import (Flask, render_template_string, request, send_from_directory, 
                   flash, redirect, url_for, after_this_request)
from werkzeug.utils import secure_filename
import openpyxl
from copy import copy
from openpyxl.utils import get_column_letter

# --- Bibliotecas para Processamento de Imagem ---
from PIL import Image
from rembg import remove

# --- Configuração Geral ---
app = Flask(__name__)
app.config['SECRET_KEY'] = 'uma-chave-secreta-muito-segura-e-diferente-56789'
app.config['MAX_CONTENT_LENGTH'] = 32 * 1024 * 1024 # Aumentado para 32MB

# Diretórios temporários. O Render lida bem com o diretório /tmp.
BASE_DIR = '/tmp/multiapp'
UPLOAD_FOLDER = os.path.join(BASE_DIR, 'uploads')
PROCESSED_FOLDER = os.path.join(BASE_DIR, 'processed')
ZIPPED_FOLDER = os.path.join(BASE_DIR, 'zipped')

# ==============================================================================
# SEÇÃO DE INICIALIZAÇÃO CORRIGIDA
# ==============================================================================

# *** CORREÇÃO APLICADA AQUI ***
# Esta função garante que as pastas existam antes de cada requisição,
# evitando o "crash" na inicialização.
@app.before_request
def ensure_dirs():
    """Garante que os diretórios de trabalho existam."""
    os.makedirs(UPLOAD_FOLDER, exist_ok=True)
    os.makedirs(PROCESSED_FOLDER, exist_ok=True)
    os.makedirs(ZIPPED_FOLDER, exist_ok=True)


# ==============================================================================
# SEÇÃO 1: LÓGICA DO CONSOLIDADOR DE PLANILHAS
# ==============================================================================

def copy_cell_style(source_cell, new_cell):
    if source_cell.has_style:
        new_cell.font = copy(source_cell.font)
        new_cell.border = copy(source_cell.border)
        new_cell.fill = copy(source_cell.fill)
        new_cell.number_format = copy(source_cell.number_format)
        new_cell.protection = copy(source_cell.protection)
        new_cell.alignment = copy(source_cell.alignment)

def processar_planilha(input_path, output_filename, sheet_name_output):
    try:
        source_workbook = openpyxl.load_workbook(input_path)
        dest_workbook = openpyxl.Workbook()
        dest_sheet = dest_workbook.active
        dest_sheet.title = sheet_name_output
        current_dest_row = 1
        is_first_sheet = True
        for sheet_name in source_workbook.sheetnames:
            source_sheet = source_workbook[sheet_name]
            if source_sheet.max_row == 0: continue
            start_row = 1 if is_first_sheet else 2
            if is_first_sheet:
                for col in range(1, source_sheet.max_column + 1):
                    col_letter = get_column_letter(col)
                    dest_sheet.column_dimensions[col_letter].width = source_sheet.column_dimensions[col_letter].width
            for r_idx, row in enumerate(source_sheet.iter_rows(min_row=start_row), start=start_row):
                dest_sheet.row_dimensions[current_dest_row].height = source_sheet.row_dimensions[r_idx].height
                for c_idx, source_cell in enumerate(row, start=1):
                    new_cell = dest_sheet.cell(row=current_dest_row, column=c_idx, value=source_cell.value)
                    copy_cell_style(source_cell, new_cell)
                current_dest_row += 1
            if source_sheet.max_row >= start_row: is_first_sheet = False
        output_path = os.path.join(UPLOAD_FOLDER, output_filename)
        dest_workbook.save(output_path)
        return output_path
    except Exception:
        print(f"--- ERRO PLANILHA ---\n{traceback.format_exc()}\n----------")
        return None

HTML_TEMPLATE_PLANILHAS = """
<!DOCTYPE html><html lang="pt-BR"><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width, initial-scale=1.0"><title>Consolidador de Planilhas</title><script src="https://cdn.tailwindcss.com"></script><link rel="stylesheet" href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap"><style>body{font-family:'Inter',sans-serif}</style></head><body class="bg-gray-100 flex items-center justify-center min-h-screen p-4"><div class="bg-white p-8 rounded-xl shadow-2xl w-full max-w-lg mx-auto"><div class="text-center mb-6"><h1 class="text-3xl font-bold text-gray-800">Consolidador de Planilhas</h1><p class="text-gray-500 mt-2">Envie um arquivo .xlsx para unir todas as abas em uma só.</p></div>{% with messages = get_flashed_messages(with_categories=true) %}{% if messages %}{% for category, message in messages %}<div class="p-4 mb-4 text-sm rounded-lg {% if category == 'error' %} bg-red-100 border border-red-400 text-red-700 {% else %} bg-green-100 border border-green-400 text-green-700 {% endif %}" role="alert"><span class="font-medium">{% if category == 'error' %}Erro:{% else %}Aviso:{% endif %}</span> {{ message }}</div>{% endfor %}{% endif %}{% endwith %}<form method="post" enctype="multipart/form-data" action="/"><div class="space-y-6"><div><label for="file-upload" class="block text-sm font-medium text-gray-700 mb-1">1. Selecione a planilha:</label><input id="file-upload" name="file" type="file" required accept=".xlsx" class="block w-full text-sm text-gray-500 file:mr-4 file:py-2 file:px-4 file:rounded-full file:border-0 file:font-semibold file:bg-indigo-50 file:text-indigo-700 hover:file:bg-indigo-100 cursor-pointer"/></div><div><label for="sheet-name" class="block text-sm font-medium text-gray-700 mb-1">2. Nomeie a aba consolidada:</label><input id="sheet-name" name="sheet_name" type="text" value="Dados_Consolidados" required class="mt-1 block w-full px-3 py-2 bg-white border border-gray-300 rounded-md shadow-sm focus:outline-none focus:ring-indigo-500 focus:border-indigo-500"></div><div><button type="submit" class="w-full flex justify-center py-3 px-4 border border-transparent rounded-md shadow-sm text-sm font-medium text-white bg-indigo-600 hover:bg-indigo-700 focus:outline-none focus:ring-2 focus:ring-offset-2 focus:ring-indigo-500">Processar e Baixar</button></div></div></form><div class="text-center mt-8"><a href="/imagens" class="text-sm text-indigo-600 hover:text-indigo-800">Ir para o Processador de Imagens &rarr;</a></div></div></body></html>
"""

# ==============================================================================
# SEÇÃO 2: LÓGICA DO PROCESSADOR DE IMAGENS
# ==============================================================================

def process_image_white_bg(input_path, output_path):
    try:
        with Image.open(input_path) as img:
            img_no_bg = remove(img)
            background = Image.new("RGBA", img_no_bg.size, (255, 255, 255, 255))
            background.paste(img_no_bg, (0, 0), img_no_bg)
            final_img = background.convert("RGB")
            final_img.save(output_path, 'JPEG', quality=95)
    except Exception:
        print(f"--- ERRO IMAGEM ---\n{traceback.format_exc()}\n----------")

HTML_TEMPLATE_IMAGENS = """
<!DOCTYPE html><html lang="pt-br"><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width, initial-scale=1.0"><title>Removedor de Fundo de Imagem</title><script src="https://cdn.tailwindcss.com"></script><link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap" rel="stylesheet"><style>body{font-family:'Inter',sans-serif}</style></head><body class="bg-gray-100 flex items-center justify-center min-h-screen"><div class="container mx-auto p-4 md:p-8 max-w-2xl"><div class="bg-white rounded-2xl shadow-lg p-8"><div class="text-center mb-8"><h1 class="text-3xl font-bold text-gray-800">Remover Fundo de Imagens</h1><p class="text-gray-500 mt-2">Envie imagens para remover o fundo (fica branco) e baixe em .zip</p></div><form action="/imagens" method="post" enctype="multipart/form-data"><div class="border-2 border-dashed border-gray-300 rounded-lg p-8 text-center cursor-pointer hover:border-blue-500 hover:bg-gray-50 transition-colors" id="dropzone"><input type="file" name="images" id="file-input" multiple required class="hidden" accept="image/*"><div id="upload-prompt"><svg class="mx-auto h-12 w-12 text-gray-400" stroke="currentColor" fill="none" viewBox="0 0 48 48" aria-hidden="true"><path d="M28 8H12a4 4 0 00-4 4v20m32-12v8m0 0v8a4 4 0 01-4 4H12a4 4 0 01-4-4v-4m32-4l-3.172-3.172a4 4 0 00-5.656 0L28 28M8 32l9.172-9.172a4 4 0 015.656 0L28 28m0 0l4 4m4-24h8m-4-4v8" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"></path></svg><p class="mt-4 text-sm text-gray-600"><span class="font-semibold text-blue-600">Clique para selecionar</span> ou arraste e solte</p></div><div id="file-list" class="mt-4 text-left text-sm text-gray-700"></div></div><div class="mt-8"><button type="submit" class="w-full bg-blue-600 text-white font-bold py-3 px-4 rounded-lg hover:bg-blue-700 disabled:bg-gray-400 transition-all" id="submit-button" disabled>Processar e Baixar</button></div></form><div class="text-center mt-8"><a href="/" class="text-sm text-blue-600 hover:text-blue-800">&larr; Voltar para o Consolidador de Planilhas</a></div></div></div><script>const dropzone=document.getElementById("dropzone"),fileInput=document.getElementById("file-input"),fileList=document.getElementById("file-list"),submitButton=document.getElementById("submit-button"),uploadPrompt=document.getElementById("upload-prompt");dropzone.onclick=()=>fileInput.click(),dropzone.ondragover=e=>{e.preventDefault(),dropzone.classList.add("border-blue-500","bg-gray-50")},dropzone.ondragleave=()=>{dropzone.classList.remove("border-blue-500","bg-gray-50")},dropzone.ondrop=e=>{e.preventDefault(),dropzone.classList.remove("border-blue-500","bg-gray-50"),e.dataTransfer.files.length&&(fileInput.files=e.dataTransfer.files,updateFileList())},fileInput.onchange=updateFileList;function updateFileList(){if(fileList.innerHTML="",fileInput.files.length){uploadPrompt.style.display="none";const e=document.createElement("ul");e.className="list-disc pl-5";for(const t of fileInput.files)e.innerHTML+=`<li>${t.name} (${(t.size/1024).toFixed(2)} KB)</li>`;fileList.appendChild(e),submitButton.disabled=!1}else uploadPrompt.style.display="block",submitButton.disabled=!0}</script></body></html>
"""

# ==============================================================================
# SEÇÃO 3: ROTAS DO APLICATIVO WEB
# ==============================================================================

@app.route('/', methods=['GET', 'POST'])
def index_planilhas():
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('Nenhum arquivo enviado.', 'error'); return redirect(request.url)
        file = request.files['file']
        if file.filename == '' or not file.filename.lower().endswith('.xlsx'):
            flash('Nenhum arquivo válido. Envie um .xlsx.', 'error'); return redirect(request.url)
        original_filename = secure_filename(file.filename)
        input_path = os.path.join(UPLOAD_FOLDER, original_filename)
        file.save(input_path)
        sheet_name = request.form.get('sheet_name', 'Dados_Consolidados')
        output_filename = f"consolidado_{original_filename}"
        output_path = processar_planilha(input_path, output_filename, sheet_name)
        if os.path.exists(input_path): os.remove(input_path)
        if output_path:
            return send_from_directory(UPLOAD_FOLDER, output_filename, as_attachment=True)
        else:
            flash('Ocorreu um erro ao processar a planilha.', 'error'); return redirect(request.url)
    return render_template_string(HTML_TEMPLATE_PLANILHAS)

@app.route('/imagens', methods=['GET', 'POST'])
def index_imagens():
    if request.method == 'POST':
        uploaded_files = request.files.getlist('images')
        if not uploaded_files or uploaded_files[0].filename == '': return "Nenhum arquivo selecionado.", 400
        session_id = str(uuid.uuid4())
        session_upload_path = os.path.join(UPLOAD_FOLDER, session_id)
        session_processed_path = os.path.join(PROCESSED_FOLDER, session_id)
        os.makedirs(session_upload_path, exist_ok=True)
        os.makedirs(session_processed_path, exist_ok=True)
        for file in uploaded_files:
            if file and file.filename:
                filename = secure_filename(file.filename)
                file.save(os.path.join(session_upload_path, filename))
        processed_filenames = []
        for filename in os.listdir(session_upload_path):
            input_path = os.path.join(session_upload_path, filename)
            new_filename = f"{os.path.splitext(filename)[0]}_sem_fundo.jpg"
            output_path = os.path.join(session_processed_path, new_filename)
            process_image_white_bg(input_path, output_path)
            processed_filenames.append(new_filename)
        zip_filename = f"imagens_sem_fundo_{session_id}.zip"
        zip_path = os.path.join(ZIPPED_FOLDER, zip_filename)
        with zipfile.ZipFile(zip_path, 'w') as zipf:
            for filename in processed_filenames:
                file_path = os.path.join(session_processed_path, filename)
                if os.path.exists(file_path):
                    zipf.write(file_path, arcname=filename)
        @after_this_request
        def cleanup(response):
            try:
                shutil.rmtree(session_upload_path)
                shutil.rmtree(session_processed_path)
                if os.path.exists(zip_path): os.remove(zip_path)
            except Exception as e:
                print(f"Erro na limpeza da sessão {session_id}: {e}")
            return response
        return send_from_directory(ZIPPED_FOLDER, zip_filename, as_attachment=True)
    return render_template_string(HTML_TEMPLATE_IMAGENS)

@app.errorhandler(413)
def request_entity_too_large(error):
    flash('O arquivo enviado é muito grande. O limite é de 32 MB.', 'error')
    if '/imagens' in request.referrer: return redirect(url_for('index_imagens'))
    return redirect(url_for('index_planilhas'))

